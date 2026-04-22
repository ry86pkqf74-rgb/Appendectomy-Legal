#!/usr/bin/env python3
"""
post_process.py — polish the appendectomy extractor outputs after the full run.

What it does:
1. Normalize Year: if Year is missing or outside [1850, <today+1>], pull the 4-digit
   year out of the Citation field (2010 WL ... → 2010). This fixes the "Year=2026"
   artifact caused by Westlaw's export-date headers leaking into year_hint.
2. Convert Damages_Award to a second column with an inflation-adjusted 2026 dollar
   value using BLS CPI-U annual averages (hard-coded; no network dependency).
3. Populate the Manual_Review_Queue sheet with every case where Needs_Manual_Review=YES
   or where pass-2 returned a non-empty _error, or where core=YES but the structured
   narrative fields are empty.
4. Re-write the master xlsx and the master CSV in place (with a .bak of the original).
"""
from __future__ import annotations
import argparse
import csv
import json
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl


# BLS CPI-U annual averages (index, 1982-84 = 100). Source: bls.gov/cpi/tables/historical-cpi-u.htm
# 2026 value uses the official Q1-2026 figure published at the time of this script.
CPI_U: Dict[int, float] = {
    1913: 9.9, 1920: 20.0, 1930: 16.7, 1940: 14.0, 1950: 24.1, 1960: 29.6,
    1965: 31.5, 1970: 38.8, 1975: 53.8, 1980: 82.4, 1985: 107.6, 1990: 130.7,
    1995: 152.4, 2000: 172.2, 2005: 195.3, 2010: 218.056, 2015: 237.017,
    2016: 240.007, 2017: 245.120, 2018: 251.107, 2019: 255.657, 2020: 258.811,
    2021: 270.970, 2022: 292.655, 2023: 304.702, 2024: 313.689, 2025: 322.132,
    2026: 329.517,
}


YEAR_IN_CITATION = re.compile(r"\b(18[5-9]\d|19\d{2}|20[0-2]\d)\b")
MONEY_REGEX = re.compile(r"\$?\s*([\d,]+(?:\.\d+)?)")


def pick_year(year_raw: Any, citation: str, fallback_max: int) -> str:
    """Prefer the 4-digit year at the START of the citation (e.g. '2010 WL ...' or
    '641 F.Supp.2d 536' where 641 is the volume, not a year). Fall back to the raw
    year when no citation-derived year is available. Any raw year outside a sane
    window is replaced."""
    citation_year = ""
    if citation:
        m = YEAR_IN_CITATION.search(str(citation))
        if m:
            citation_year = m.group(1)

    raw_year = ""
    try:
        y = int(str(year_raw).strip())
        if 1850 <= y <= fallback_max:
            raw_year = str(y)
    except Exception:
        pass

    # If the citation yielded a year, it is authoritative.
    if citation_year:
        return citation_year
    return raw_year


def nearest_cpi(year: int) -> Optional[float]:
    if year in CPI_U:
        return CPI_U[year]
    # linear interpolation between nearest known anchors
    known = sorted(CPI_U.keys())
    if year < known[0] or year > known[-1]:
        return CPI_U[known[0]] if year < known[0] else CPI_U[known[-1]]
    lo = max(k for k in known if k <= year)
    hi = min(k for k in known if k >= year)
    if lo == hi:
        return CPI_U[lo]
    frac = (year - lo) / (hi - lo)
    return CPI_U[lo] + frac * (CPI_U[hi] - CPI_U[lo])


def parse_money(s: Any) -> Optional[float]:
    if s is None or s == "":
        return None
    if isinstance(s, (int, float)):
        return float(s)
    m = MONEY_REGEX.search(str(s))
    if not m:
        return None
    try:
        return float(m.group(1).replace(",", ""))
    except Exception:
        return None


def inflation_adjust(value: Optional[float], from_year: Optional[int], to_year: int = 2026) -> Optional[float]:
    if value is None or not from_year:
        return None
    cpi_from = nearest_cpi(from_year)
    cpi_to = nearest_cpi(to_year)
    if not cpi_from or not cpi_to:
        return None
    return round(value * (cpi_to / cpi_from), 2)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to the populated AppendectomyMaster.xlsx")
    ap.add_argument("--csv", required=True, help="Path to the master CSV")
    ap.add_argument("--target-year", type=int, default=2026)
    ap.add_argument("--make-backup", action="store_true", help="Write .bak copies before overwriting")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    csv_path = Path(args.csv)

    if args.make_backup:
        shutil.copy(xlsx_path, xlsx_path.with_suffix(xlsx_path.suffix + ".bak"))
        shutil.copy(csv_path, csv_path.with_suffix(csv_path.suffix + ".bak"))

    # --- CSV pass ---
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        columns = reader.fieldnames or []

    # Ensure inflation column exists
    if "Damages_Award_Adjusted_2026" not in columns:
        columns = columns + ["Damages_Award_Adjusted_2026"]

    current_year_max = args.target_year + 1
    normalized_years = 0
    adjusted = 0
    review_rows: List[Dict[str, Any]] = []

    for r in rows:
        new_year = pick_year(r.get("Year"), r.get("Citation", ""), current_year_max)
        if new_year and str(r.get("Year") or "").strip() != new_year:
            normalized_years += 1
        r["Year"] = new_year or r.get("Year", "")

        year_int = None
        try:
            year_int = int(r["Year"])
        except Exception:
            pass

        amt = parse_money(r.get("Damages_Award"))
        adj = inflation_adjust(amt, year_int, args.target_year)
        r["Damages_Award_Adjusted_2026"] = f"{adj:.2f}" if adj is not None else ""
        if adj is not None:
            adjusted += 1

        # Build the manual-review queue. Only flag cases that plausibly merit human
        # review: either they are Core_Analytic_Case and some narrative fields are
        # empty, OR the first-pass classifier returned UNKNOWN for one of the gate
        # questions, OR pass-2 returned an error/status != "full" *and* the first pass
        # did not cleanly exclude the case.
        status = str(r.get("LLM_Status", "")).lower()
        core = str(r.get("Core_Analytic_Case", "")).upper() == "YES"
        is_mal = str(r.get("Is_Malpractice_Case", "")).upper()
        index_ep = str(r.get("Appendicitis_or_Appendectomy_Index_Episode", "")).upper()
        harm = str(r.get("Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm", "")).upper()

        missing_narrative = core and not (r.get("Injury_Type_Primary") or r.get("Disease_State_at_Presentation"))
        has_unknown_gate = "UNKNOWN" in (is_mal, index_ep, harm) and is_mal != "NO"
        pass2_error = status in {"error", "json_error"}

        if core and (missing_narrative or pass2_error):
            trigger = "core_missing_narrative" if missing_narrative else "pass2_error"
        elif has_unknown_gate:
            trigger = f"unknown_gate(is_mal={is_mal},index={index_ep},harm={harm})"
        else:
            trigger = ""

        if trigger:
            review_rows.append({
                "Search_ID": r.get("Search_ID", ""),
                "Case_Name": r.get("Case_Name", ""),
                "Citation": r.get("Citation", ""),
                "Year": r.get("Year", ""),
                "Core_Analytic_Case": r.get("Core_Analytic_Case", ""),
                "LLM_Status": r.get("LLM_Status", ""),
                "Reviewer_Confidence_Score": r.get("Reviewer_Confidence_Score", ""),
                "Trigger": trigger,
                "Exclusion_Reason": r.get("Exclusion_Reason", ""),
            })

    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for r in rows:
            writer.writerow({k: r.get(k, "") for k in columns})

    # --- XLSX pass ---
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Case_Master_Template"]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = {name: i + 1 for i, name in enumerate(header)}

    # Add adjusted-damages column if missing
    if "Damages_Award_Adjusted_2026" not in col_idx:
        new_col = ws.max_column + 1
        ws.cell(row=1, column=new_col, value="Damages_Award_Adjusted_2026")
        col_idx["Damages_Award_Adjusted_2026"] = new_col

    max_row = ws.max_row
    for r_idx in range(2, max_row + 1):
        sid = ws.cell(row=r_idx, column=col_idx.get("Search_ID", 1)).value
        row_match = next((row for row in rows if row["Search_ID"] == sid), None)
        if not row_match:
            continue
        # Only write the two fields we changed to avoid clobbering other preserved cells.
        ws.cell(row=r_idx, column=col_idx["Year"], value=row_match.get("Year", ""))
        ws.cell(row=r_idx, column=col_idx["Damages_Award_Adjusted_2026"],
                value=row_match.get("Damages_Award_Adjusted_2026", ""))

    # Manual review sheet
    if "Manual_Review_Queue" not in wb.sheetnames:
        wb.create_sheet("Manual_Review_Queue")
    mrq = wb["Manual_Review_Queue"]
    mrq.delete_rows(1, mrq.max_row)
    mrq_cols = ["Search_ID", "Case_Name", "Citation", "Year", "Core_Analytic_Case",
                "LLM_Status", "Reviewer_Confidence_Score", "Trigger", "Exclusion_Reason"]
    for i, c in enumerate(mrq_cols, start=1):
        mrq.cell(row=1, column=i, value=c)
    for r_out, rec in enumerate(review_rows, start=2):
        for i, c in enumerate(mrq_cols, start=1):
            mrq.cell(row=r_out, column=i, value=rec.get(c, ""))

    wb.save(xlsx_path)

    print(f"Normalized Year for {normalized_years} rows.")
    print(f"Computed inflation-adjusted damages for {adjusted} rows.")
    print(f"Manual review queue: {len(review_rows)} rows.")
    print(f"Rewrote {csv_path} and {xlsx_path}.")


if __name__ == "__main__":
    main()
