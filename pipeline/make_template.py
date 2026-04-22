#!/usr/bin/env python3
"""
make_template.py — generate the Template.xlsx for appendectomy extraction.

This is the appendectomy analogue of the chole project's Template.xlsx. It contains:
  - Read_Me              — study documentation
  - Case_Master_Template — one row per case (full schema)
  - Data_Dictionary      — column name, definition, allowed values, source notes
  - Extended_Extraction  — narrative/long-form companion sheet
  - Manual_Review_Queue  — auto-populated from post-processing

Run once after `pip install openpyxl`:
    python make_template.py --out Template.xlsx
"""
from __future__ import annotations
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


# -----------------------------------------------------------------------------
# Master schema for the Case_Master_Template sheet.
# Order matches what the extractor expects.
# -----------------------------------------------------------------------------
TEMPLATE_COLUMNS = [
    # --- Identifiers / search metadata ---
    "Search_ID", "Search_Group", "Search_Term", "Source_Database", "File_Name",
    # --- Case identity ---
    "Case_Name", "Citation", "Year", "Court", "Jurisdiction",
    # --- Triage / relevance (appendectomy specific) ---
    "Is_Malpractice_Case",
    "Appendicitis_or_Appendectomy_Index_Episode",
    "Index_Procedure_Appendectomy",
    "Appendicitis_Diagnosis_Case",
    "Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm",
    "Core_Analytic_Case", "Exclusion_Reason", "Duplicate_Group_ID",
    # --- Legal posture / outcome ---
    "Legal_Case_Type", "Procedural_Posture", "Legal_Outcome",
    "Damages_Award", "Settlement_Amount",
    "Economic_Damages", "NonEconomic_Damages", "Punitive_Damages",
    "Time_to_Resolution_Years",
    "Appellate_Status",
    "Expert_Testimony_Mentioned", "Expert_Testimony_Type",
    "Expert_Criticism_Text", "Defense_Strategy_Summary",
    "Alleged_Breach_Categories",
    # --- Procedure / disease state ---
    "Index_Procedure_Type",       # urgent-emergent / interval / incidental / no appendectomy / unclear
    "Procedure_Approach",         # laparoscopic / converted / open / robotic / unclear
    "Disease_State_at_Presentation",  # uncomplicated / perforated / abscess / gangrenous / chronic / unclear
    "Injury_Type_Primary",        # delayed diagnosis w/ perforation / failed appendectomy / stump / bowel / bleeding / abscess / leak / obstruction / fertility / mixed / other
    "Injury_Type_Secondary",
    "Injury_Severity",            # major / minor / unknown
    "Wrong_Structure_Removed",
    "Appendix_Not_Removed",
    "Need_for_Reoperation",
    "Need_for_Bowel_Resection",
    "Need_for_Stoma",
    "Tertiary_Referral",
    "Death",
    "Long_Term_Morbidity",
    # --- Timing / management ---
    "Recognition_Timing",         # preoperative delayed dx / intraoperative / early postop / delayed post-discharge / unknown
    "Recognition_Timing_Detail",
    "Time_From_Presentation_To_Diagnosis_Hours",
    "Time_From_Surgery_To_Recognition_Days",
    "Delay_Days",
    "Delayed_Diagnosis_Alleged",
    "Improper_Postop_Management_Alleged",
    "Failure_to_Refer_Alleged",
    "NonSpecialist_Repair_or_Management",
    # --- Snippet fields (short verbatim) ---
    "Operative_Text_Snippet",
    "Difficulty_Text_Snippet",
    "Recognition_Text_Snippet",
    # --- Difficulty / operative complexity ---
    "Difficulty_Assessability",   # clear | possible | not assessable
    "Difficult_Case",
    "Perforated_or_Gangrenous_Appendix",
    "Abscess_or_Phlegmon",
    "Severe_Inflammation",
    "Dense_Adhesions",
    "Obesity_or_Habitus_Difficulty",
    "Retrocecal_or_Unusual_Appendix_Location",
    "Difficult_Dissection",
    "Bleeding_Obscuring_Field",
    "Conversion_to_Open",
    "Bowel_Resection_or_Ileocecectomy",
    "Stump_Leak_or_Stump_Problem",
    "Appendix_Not_Removed_or_Wrong_Tissue",
    "Difficulty_Documented",     # explicit | inferred | not documented
    "Difficulty_Recognized_By_Surgeon",
    "Adaptation_Performed",
    "Adaptation_Type",           # conversion / drain / bowel resection / interval / antibiotics-first / subtotal / call-for-help / referral / aborted / other / none
    "Adaptation_Appears_Appropriate",
    # --- Other medicolegal ---
    "Problematic_Visualization_Alleged",
    "Aberrant_Anatomy_Mentioned",
    "Inadequate_Informed_Consent_Alleged",
    "Poor_Communication_Alleged",
    "Unexpected_Postop_Course_Referenced",
    "Plaintiff_Demographics",
    "Surgeon_Characteristics",
    "Facility_Type",
    "Guideline_Adherence_Mentioned",
    "Total_Healthcare_Cost_Mentioned",
    "Preventability_Assessment",
    "Reviewer_Confidence_Score",
    "Reviewer_Notes",
    # --- First-pass classifier fields (kept for every row) ---
    "First_Pass_Likely_Case_Type",
    "First_Pass_Rationale",
    "Full_Extraction_Performed",
    "LLM_Status",
    "Needs_Manual_Review",
]


EXTENDED_COLUMNS = [
    "Search_ID",
    "Comorbid_Diagnoses_Text",
    "Operative_Findings_Detail",
    "Plaintiff_Claims_Expanded",
    "Plaintiff_Medical_Support_Summary",
    "Defense_Medical_Rebuttal_Summary",
    "Plaintiff_Expert_Summary",
    "Defense_Expert_Summary",
    "Court_Medical_Reasoning_Summary",
    "Claim_Support_Matrix_JSON",
    "Evidence_Quotes_JSON",
    "Extended_Extraction_Notes",
    "LLM_Status",
]


REVIEW_COLUMNS = [
    "Search_ID", "Case_Name", "Citation",
    "Reviewer_Confidence_Score", "Core_Analytic_Case",
    "Reason_For_Review", "Reviewer_Notes",
]


# (column_name, definition, allowed_values, source_notes)
DATA_DICT = [
    ("Search_ID", "Stable per-case ID = <file_stub>_<segment_index>.", "Free text", "Auto-generated"),
    ("Search_Group", "Logical group of searches (set per Westlaw query).", "Free text", "CLI flag"),
    ("Search_Term", "Original Westlaw query string.", "Free text", "CLI flag (cannot be recovered from RTF)"),
    ("Source_Database", "Origin database for the export.", "Westlaw", "Constant"),
    ("File_Name", "Source RTF file the case came from.", "Free text", "Auto"),
    ("Case_Name", "Full caption (Plaintiff v. Defendant).", "Free text", "LLM + regex fallback"),
    ("Citation", "First Westlaw / reporter citation found.", "Free text", "LLM + regex fallback"),
    ("Year", "4-digit year of decision.", "YYYY", "LLM + regex fallback"),
    ("Court", "Issuing court as printed in the opinion.", "Free text", "LLM + regex fallback"),
    ("Jurisdiction", "Federal / State / Unknown.", "Federal | State | Unknown", "Inferred from court"),
    ("Is_Malpractice_Case", "Medical malpractice/negligence/informed-consent/standard-of-care suit.", "YES | NO | UNKNOWN", "LLM"),
    ("Appendicitis_or_Appendectomy_Index_Episode",
     "The lawsuit centers on suspected appendicitis, appendectomy, failure to dx appendicitis, delayed surgery, or appendectomy-related postop management.",
     "YES | NO | UNKNOWN", "LLM"),
    ("Index_Procedure_Appendectomy",
     "Appendectomy (or attempted appendectomy) was the surgery at issue — not remote history.",
     "YES | NO | UNKNOWN", "LLM"),
    ("Appendicitis_Diagnosis_Case",
     "Case hinges on a diagnostic-timing issue for appendicitis (delayed/missed dx, failure to operate).",
     "YES | NO | UNKNOWN", "LLM"),
    ("Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm",
     "A material injury/complication/consequence tied to the episode: perforation, abscess, sepsis, fertility injury, bowel injury, reoperation, failed appendectomy, stump appendicitis, prolonged hospitalization, death.",
     "YES | NO | UNKNOWN", "LLM"),
    ("Core_Analytic_Case", "All three screening fields = YES.", "YES | NO", "Auto-derived"),
    ("Exclusion_Reason", "Why the case is not core analytic.", "Free text", "Auto-derived"),
    ("Duplicate_Group_ID", "Tag for cases with identical normalized Case_Name+citation.", "DUP-xxxxx | blank", "Auto-derived"),
    ("Legal_Case_Type", "Document type (Opinion, Settlement summary, etc.).", "Free text", "LLM"),
    ("Procedural_Posture", "Trial verdict / Appeal / SJ / etc.", "Free text", "LLM"),
    ("Legal_Outcome", "Who prevailed.", "Plaintiff-favorable | Defense-favorable | Settlement | Mixed | Unknown", "LLM"),
    ("Damages_Award", "Dollar amount of damages awarded — only if explicit.", "$ amount | blank", "LLM"),
    ("Settlement_Amount", "Dollar settlement — only if explicit.", "$ amount | blank", "LLM"),
    ("Economic_Damages", "Economic component if itemized.", "$ amount | blank", "LLM"),
    ("NonEconomic_Damages", "Non-economic component if itemized.", "$ amount | blank", "LLM"),
    ("Punitive_Damages", "Punitive component if itemized.", "$ amount | blank", "LLM"),
    ("Time_to_Resolution_Years", "Surgery -> final resolution, years.", "Numeric | blank", "LLM"),
    ("Appellate_Status", "Was the case appealed; if so, outcome.",
     "Original | Appeal - Plaintiff win | Appeal - Defense win | Appeal - Remanded | Appeal - Mixed | Unknown", "LLM"),
    ("Expert_Testimony_Mentioned", "Any expert testimony mentioned.", "YES | NO", "LLM"),
    ("Expert_Testimony_Type", "Side(s) presenting expert testimony.",
     "Plaintiff only | Defense only | Both | None mentioned | Unknown", "LLM"),
    ("Expert_Criticism_Text", "Concise summary of expert criticism (<=500 chars).", "Free text", "LLM"),
    ("Defense_Strategy_Summary", "Defense theory in 1-3 sentences.", "Free text", "LLM"),
    ("Alleged_Breach_Categories",
     "Comma-list of alleged breaches (see README). e.g. 'delayed diagnosis of appendicitis, failure to order imaging'.",
     "Free text (controlled vocab)", "LLM"),
    ("Index_Procedure_Type", "Urgency/setting of the appendectomy if done.",
     "urgent-emergent appendectomy | interval appendectomy | incidental appendectomy | no appendectomy performed | unclear", "LLM"),
    ("Procedure_Approach", "Approach used.",
     "laparoscopic | converted | open | robotic | unclear", "LLM"),
    ("Disease_State_at_Presentation", "Appendix condition at presentation/operation.",
     "uncomplicated appendicitis | perforated appendicitis | abscess-phlegmon | gangrenous-necrotic appendicitis | chronic-recurrent appendicitis | unclear", "LLM"),
    ("Injury_Type_Primary", "Primary injury/complication alleged.",
     "delayed diagnosis with perforation | failed appendectomy | stump appendicitis | bowel injury | bleeding-vascular injury | abscess-infection | leak | obstruction-adhesive complication | fertility injury | mixed | other", "LLM"),
    ("Injury_Type_Secondary", "Any secondary injury.", "Free text", "LLM"),
    ("Injury_Severity", "Major vs Minor.", "major | minor | unknown", "LLM"),
    ("Wrong_Structure_Removed", "Wrong tissue removed.", "YES | NO | UNKNOWN", "LLM"),
    ("Appendix_Not_Removed", "Appendix left in situ when it should have been removed.", "YES | NO | UNKNOWN", "LLM"),
    ("Need_for_Reoperation", "Required reoperation.", "YES | NO | UNKNOWN", "LLM"),
    ("Need_for_Bowel_Resection", "Required bowel resection.", "YES | NO | UNKNOWN", "LLM"),
    ("Need_for_Stoma", "Required ostomy.", "YES | NO | UNKNOWN", "LLM"),
    ("Tertiary_Referral", "Referred to tertiary care.", "YES | NO | UNKNOWN", "LLM"),
    ("Death", "Patient died.", "YES | NO | UNKNOWN", "LLM"),
    ("Long_Term_Morbidity", "Documented long-term morbidity.", "YES | NO | UNKNOWN", "LLM"),
    ("Recognition_Timing", "When the complication/dx was recognized.",
     "preoperative delayed diagnosis | intraoperative | early postoperative | delayed after discharge | unknown", "LLM"),
    ("Recognition_Timing_Detail", "Short factual detail.", "Free text", "LLM"),
    ("Time_From_Presentation_To_Diagnosis_Hours", "Hours from ED/clinic to appendicitis dx (if stated).", "Numeric | blank", "LLM"),
    ("Time_From_Surgery_To_Recognition_Days", "Days post-op to recognition of complication.", "Numeric | blank", "LLM"),
    ("Delay_Days", "Overall diagnostic/operative delay reported in the opinion.", "Numeric | blank", "LLM"),
    ("Delayed_Diagnosis_Alleged", "Delayed dx specifically alleged.", "YES | NO", "LLM"),
    ("Improper_Postop_Management_Alleged", "Postop management alleged improper.", "YES | NO", "LLM"),
    ("Failure_to_Refer_Alleged", "Failure to refer/transfer alleged.", "YES | NO", "LLM"),
    ("NonSpecialist_Repair_or_Management", "Non-specialist managed the complication.", "YES | NO | UNKNOWN", "LLM"),
    ("Operative_Text_Snippet", "Verbatim short quote of operative description.", "Free text (<=400 chars)", "LLM"),
    ("Difficulty_Text_Snippet", "Verbatim snippet describing difficulty/complexity.", "Free text (<=400 chars)", "LLM"),
    ("Recognition_Text_Snippet", "Verbatim snippet on when/how complication was recognized.", "Free text (<=400 chars)", "LLM"),
    ("Difficulty_Assessability", "Can operative difficulty be assessed from the opinion?",
     "clear | possible | not assessable", "LLM"),
    ("Difficult_Case", "Operative difficulty documented.", "YES | NO | UNKNOWN", "LLM"),
    ("Perforated_or_Gangrenous_Appendix", "Perforation or gangrene documented.", "YES | NO | UNKNOWN", "LLM"),
    ("Abscess_or_Phlegmon", "Abscess or phlegmon documented.", "YES | NO | UNKNOWN", "LLM"),
    ("Severe_Inflammation", "Severe inflammation documented.", "YES | NO | UNKNOWN", "LLM"),
    ("Dense_Adhesions", "Dense adhesions documented.", "YES | NO | UNKNOWN", "LLM"),
    ("Obesity_or_Habitus_Difficulty", "Obesity/habitus difficulty mentioned.", "YES | NO | UNKNOWN", "LLM"),
    ("Retrocecal_or_Unusual_Appendix_Location", "Retrocecal or unusual location documented.", "YES | NO | UNKNOWN", "LLM"),
    ("Difficult_Dissection", "Difficult dissection documented.", "YES | NO | UNKNOWN", "LLM"),
    ("Bleeding_Obscuring_Field", "Bleeding obscured the operative field.", "YES | NO | UNKNOWN", "LLM"),
    ("Conversion_to_Open", "Laparoscopic converted to open.", "YES | NO | UNKNOWN", "LLM"),
    ("Bowel_Resection_or_Ileocecectomy", "Bowel resection or ileocecectomy performed.", "YES | NO | UNKNOWN", "LLM"),
    ("Stump_Leak_or_Stump_Problem", "Stump leak/stump appendicitis/stump problem.", "YES | NO | UNKNOWN", "LLM"),
    ("Appendix_Not_Removed_or_Wrong_Tissue", "Appendix not removed or wrong tissue removed.", "YES | NO | UNKNOWN", "LLM"),
    ("Difficulty_Documented", "How difficulty is captured in the record.",
     "explicit | inferred | not documented", "LLM"),
    ("Difficulty_Recognized_By_Surgeon", "Surgeon recognized the difficulty.", "YES | NO | UNKNOWN", "LLM"),
    ("Adaptation_Performed", "Intraoperative adaptation performed.", "YES | NO | UNKNOWN", "LLM"),
    ("Adaptation_Type", "Type of adaptation.",
     "conversion | drain | bowel resection | interval management | antibiotics-first | subtotal-partial | call for help | referral | aborted | other | none", "LLM"),
    ("Adaptation_Appears_Appropriate", "Adaptation appears appropriate given findings.", "YES | NO | UNKNOWN", "LLM"),
    ("Problematic_Visualization_Alleged", "Visualization problems alleged.", "YES | NO", "LLM"),
    ("Aberrant_Anatomy_Mentioned", "Aberrant anatomy mentioned.", "YES | NO | UNKNOWN", "LLM"),
    ("Inadequate_Informed_Consent_Alleged", "Informed-consent claim.", "YES | NO", "LLM"),
    ("Poor_Communication_Alleged", "Poor communication alleged.", "YES | NO", "LLM"),
    ("Unexpected_Postop_Course_Referenced", "Unexpected postop course referenced.", "YES | NO | UNKNOWN", "LLM"),
    ("Plaintiff_Demographics", "Age/sex/relevant demographics.", "Free text", "LLM"),
    ("Surgeon_Characteristics", "Surgeon specialty/experience if stated.", "Free text", "LLM"),
    ("Facility_Type", "Facility type if stated.", "Free text", "LLM"),
    ("Guideline_Adherence_Mentioned", "Guideline adherence discussed.", "YES | NO | UNKNOWN", "LLM"),
    ("Total_Healthcare_Cost_Mentioned", "Healthcare cost figure mentioned.", "YES | NO", "LLM"),
    ("Preventability_Assessment", "Court/expert preventability statement summary.", "Free text", "LLM"),
    ("Reviewer_Confidence_Score", "Model's self-reported confidence 1-5.", "1-5", "LLM"),
    ("Reviewer_Notes", "Free-form notes.", "Free text", "LLM"),
    ("First_Pass_Likely_Case_Type", "First-pass classifier label.",
     "delayed_diagnosis | operative_complication | postop_management | payment_dispute | incidental_history | unclear", "LLM pass 1"),
    ("First_Pass_Rationale", "Short (<=200 char) rationale from pass 1.", "Free text", "LLM pass 1"),
    ("Full_Extraction_Performed", "Was pass-2 extraction run on this case?", "YES | NO", "Auto"),
    ("LLM_Status", "Completion status of the LLM call.",
     "full | mock | skipped | error | first_pass_only", "Auto"),
    ("Needs_Manual_Review", "Flagged for manual review.", "YES | NO", "Auto"),
]


READ_ME_LINES = [
    "Appendectomy Medicolegal Extraction — Westlaw Corpus",
    "",
    "Purpose",
    "This workbook is the structured extraction template for an academic study of appendectomy-",
    "and appendicitis-related malpractice litigation. Each row in Case_Master_Template represents",
    "one Westlaw case. The Extended_Extraction sheet holds richer free-text fields keyed by Search_ID.",
    "",
    "Key principle: do not force values",
    "If the source does not clearly support a value, leave the field as UNKNOWN (categorical/boolean)",
    "or blank (text/numeric). Snippet fields (Operative_Text_Snippet, Difficulty_Text_Snippet,",
    "Recognition_Text_Snippet) exist so a reviewer can verify each judgment quickly.",
    "",
    "Screening — a case is Core_Analytic_Case = YES only if ALL of these are YES:",
    "  A. Is_Malpractice_Case",
    "  B. Appendicitis_or_Appendectomy_Index_Episode",
    "  C. Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm",
    "Delayed/missed diagnosis of appendicitis qualifies as a true positive even if no appendectomy",
    "was ever performed — this is the main domain difference from the chole project.",
    "",
    "Exclude: social security / disability decisions, insurance underwriting, ERISA / billing",
    "disputes, prison civil-rights GI cases, incidental past-surgical-history mentions, and any",
    "case where appendectomy is only background.",
    "",
    "Workflow",
    "1. Run the extractor (appendectomy_extractor.py) against the RTF exports.",
    "2. Run post_process.py to populate Manual_Review_Queue and assorted derived fields.",
    "3. Filter Case_Master_Template on Core_Analytic_Case = YES.",
    "4. For each Core_Analytic_Case = YES row, review (in order):",
    "     a. Operative_Text_Snippet, Difficulty_Text_Snippet, Recognition_Text_Snippet",
    "     b. Adaptation_Type, Alleged_Breach_Categories",
    "     c. Expert_Criticism_Text, Defense_Strategy_Summary",
    "     d. Reviewer_Confidence_Score (anything <=3 deserves a manual pass)",
    "5. Review the Extended_Extraction sheet for cases where structured fields look thin.",
    "",
    "Coding conventions",
    "YES / NO / UNKNOWN     — clinical YES_NO_UNKNOWN fields",
    "YES / NO               — allegation-style fields (default NO when not alleged)",
    "Money fields           — only when an explicit dollar figure appears in the source",
    "",
    "Important caveats",
    "- Search_Term cannot be recovered from the RTF; supply it via --search-term when running.",
    "- Duplicate_Group_ID is conservative (normalized case-name match only).",
    "- Reviewer_Confidence_Score is the model's self-reported confidence.",
    "- This workbook is a decision-support extractor, not a final adjudicator.",
]


def build_workbook(out_path: Path) -> None:
    wb = openpyxl.Workbook()
    # Read_Me
    ws = wb.active
    ws.title = "Read_Me"
    for i, line in enumerate(READ_ME_LINES, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.column_dimensions["A"].width = 110

    # Case_Master_Template
    ws = wb.create_sheet("Case_Master_Template")
    bold = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for j, col in enumerate(TEMPLATE_COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"

    # Data_Dictionary
    ws = wb.create_sheet("Data_Dictionary")
    for j, h in enumerate(["Column_Name", "Definition", "Allowed_Values", "Source_Notes"], start=1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = bold
        c.fill = header_fill
    for i, (name, defn, allowed, src) in enumerate(DATA_DICT, start=2):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=defn)
        ws.cell(row=i, column=3, value=allowed)
        ws.cell(row=i, column=4, value=src)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 28
    ws.freeze_panes = "A2"

    # Extended_Extraction
    ws = wb.create_sheet("Extended_Extraction")
    for j, col in enumerate(EXTENDED_COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = bold
        c.fill = header_fill
    ws.freeze_panes = "A2"

    # Manual_Review_Queue
    ws = wb.create_sheet("Manual_Review_Queue")
    for j, col in enumerate(REVIEW_COLUMNS, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = bold
        c.fill = header_fill
    ws.cell(row=3, column=1,
            value="(Auto-populated by post_process.py — any case with confidence<=3 or LLM_Status!='full' lands here.)")
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Wrote {out_path}")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="Template.xlsx")
    args = ap.parse_args()
    build_workbook(Path(args.out))


if __name__ == "__main__":
    main()
