#!/usr/bin/env python3
"""
appendectomy_extractor.py — two-pass structured extraction over Westlaw appendectomy RTF exports.

Pipeline (single local OpenAI-compatible inference server, only gpt-oss-120b):
  1. Preprocessing
     - Parse every RTF matching --input-glob with striprtf
     - Split each file on "End of Document" delimiters
     - Clean obvious RTF/header/hyperlink artifacts
     - Emit a manifest CSV (Search_ID, File_Name, Case_Name_hint, Citation_hint, Char_count)

  2. First-pass classifier (fast, low-token)
     - For every case: YES/NO/UNKNOWN on:
         Is_Malpractice_Case
         Appendicitis_or_Appendectomy_Index_Episode
         Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm
     - Plus likely_case_type and a short rationale.
     - Only likely positives advance to pass 2, but every case keeps a row.

  3. Second-pass structured extractor (longer prompt, richer snippets)
     - One JSON object per case matching the Template.xlsx columns exactly.
     - Uses vLLM xgrammar guided-JSON output.

  4. Deduplication
     - Normalize (case_name, citation) -> Duplicate_Group_ID, flag later rows with Exclusion_Reason=duplicate.

  5. Outputs
     - xlsx master workbook (populated from Template.xlsx)
     - JSONL audit file with the full structured extraction and evidence
     - Extended CSV (second-sheet-style) for richer narrative fields
     - Exclusions CSV
     - Resumable: JSONL checkpoints are appended per case

Everything is deterministic (temperature=0) and resumable from the JSONL checkpoint file.

CLI
---
Mock dry run (no GPU, no LLM):
    python appendectomy_extractor.py \
        --input-glob "/workspace/appendectomy_data/*.rtf" \
        --template /workspace/appendectomy_pipeline/Template.xlsx \
        --output-xlsx /workspace/appendectomy_out/appendectomy_mock.xlsx \
        --mock-llm --max-cases 25

Full run against local vLLM:
    python appendectomy_extractor.py \
        --input-glob "/workspace/appendectomy_data/*.rtf" \
        --template /workspace/appendectomy_pipeline/Template.xlsx \
        --output-xlsx /workspace/appendectomy_out/appendectomy_cases_output.xlsx \
        --output-csv /workspace/appendectomy_out/appendectomy_cases_output.csv \
        --output-jsonl /workspace/appendectomy_out/appendectomy_cases_output.jsonl \
        --output-extended-csv /workspace/appendectomy_out/appendectomy_cases_extended.csv \
        --output-manifest /workspace/appendectomy_out/appendectomy_cases_manifest.csv \
        --output-exclusions /workspace/appendectomy_out/appendectomy_cases_exclusions.csv \
        --search-group "Westlaw appendectomy search" \
        --search-term "appendectomy malpractice Westlaw full-text export" \
        --base-url http://localhost:8000/v1 \
        --model gptoss120b \
        --workers 8
"""
from __future__ import annotations
import argparse
import concurrent.futures as cf
import csv
import glob
import hashlib
import json
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

try:
    import requests
except ImportError:
    requests = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from striprtf.striprtf import rtf_to_text
except ImportError:
    rtf_to_text = None


SOURCE_DATABASE = "Westlaw"
CASE_DELIMITER = re.compile(r"\n\s*End of Document\s*\n", re.IGNORECASE)
CITATION_REGEX = re.compile(r"(\d{4}\s+WL\s+\d+|\d+\s+[A-Z][A-Za-z.]+\s*\d*d?\s+\d+|\d{4}\s+U\.S\.\s+Dist\.\s+LEXIS\s+\d+)")
CASE_NAME_REGEX = re.compile(r"^\s*([A-Z][^\n]{3,120}?\sv\.\s[^\n]{2,120})\s*$", re.MULTILINE)
YEAR_REGEX = re.compile(r"\b(19|20)\d{2}\b")


# -----------------------------------------------------------------------------
# Data structures
# -----------------------------------------------------------------------------
@dataclass
class CaseChunk:
    search_id: str
    file_name: str
    segment_index: int
    text: str
    case_name_hint: str = ""
    citation_hint: str = ""
    year_hint: str = ""
    char_count: int = 0

    def to_manifest_row(self) -> Dict[str, Any]:
        return {
            "Search_ID": self.search_id,
            "File_Name": self.file_name,
            "Segment_Index": self.segment_index,
            "Case_Name_Hint": self.case_name_hint,
            "Citation_Hint": self.citation_hint,
            "Year_Hint": self.year_hint,
            "Char_Count": self.char_count,
        }


# -----------------------------------------------------------------------------
# RTF preprocessing
# -----------------------------------------------------------------------------
def load_rtf(path: Path) -> str:
    if rtf_to_text is None:
        raise RuntimeError("striprtf is not installed (pip install striprtf)")
    raw = path.read_text(encoding="utf-8", errors="ignore")
    text = rtf_to_text(raw, errors="ignore")
    return text


def clean_case_text(t: str) -> str:
    # Strip picture/hyperlink residues and normalize whitespace
    t = re.sub(r"\{\\\*[^{}]*\}", " ", t)
    t = re.sub(r"HYPERLINK\s+\"[^\"]+\"", " ", t)
    t = re.sub(r"\r", "\n", t)
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    t = unicodedata.normalize("NFKC", t)
    return t.strip()


def split_cases(file_path: Path) -> List[str]:
    text = load_rtf(file_path)
    chunks = CASE_DELIMITER.split(text)
    # Drop blank/very-small chunks (Westlaw sometimes emits trailing empty segments)
    chunks = [clean_case_text(c) for c in chunks]
    return [c for c in chunks if len(c) > 400]


def extract_hints(chunk: str) -> Tuple[str, str, str]:
    m_name = CASE_NAME_REGEX.search(chunk[:3000])
    case_name = m_name.group(1).strip() if m_name else ""
    m_cite = CITATION_REGEX.search(chunk[:5000])
    citation = m_cite.group(0).strip() if m_cite else ""
    year = ""
    m_year = YEAR_REGEX.search(chunk[:5000])
    if m_year:
        year = m_year.group(0)
    return case_name, citation, year


def build_case_chunks(rtf_paths: List[Path]) -> List[CaseChunk]:
    cases: List[CaseChunk] = []
    for p in rtf_paths:
        try:
            segments = split_cases(p)
        except Exception as e:
            print(f"[WARN] failed to parse {p.name}: {e}", file=sys.stderr)
            continue
        stub = re.sub(r"[^A-Za-z0-9]+", "_", p.stem).strip("_").lower()
        for i, seg in enumerate(segments):
            sid = f"{stub}_{i:04d}"
            case_name, cit, year = extract_hints(seg)
            cases.append(CaseChunk(
                search_id=sid,
                file_name=p.name,
                segment_index=i,
                text=seg,
                case_name_hint=case_name,
                citation_hint=cit,
                year_hint=year,
                char_count=len(seg),
            ))
    return cases


# -----------------------------------------------------------------------------
# LLM client (OpenAI-compatible)
# -----------------------------------------------------------------------------
class LLMClient:
    def __init__(self, base_url: str, model: str, timeout: int = 180, mock: bool = False):
        self.base_url = base_url.rstrip("/")
        self.model = model
        self.timeout = timeout
        self.mock = mock

    def chat(self, messages: List[Dict[str, str]], response_format: Optional[Dict] = None,
             max_tokens: int = 2048, temperature: float = 0.0) -> str:
        if self.mock:
            # Deterministic stub for dry runs without a server
            sys_msg = next((m for m in messages if m["role"] == "system"), {"content": ""})
            return self._mock_response(messages[-1]["content"], "pass2" if "EXACT OUTPUT SCHEMA" in sys_msg["content"] + messages[-1]["content"] else "pass1")

        if requests is None:
            raise RuntimeError("requests is not installed")

        payload: Dict[str, Any] = {
            "model": self.model,
            "messages": messages,
            "temperature": temperature,
            "top_p": 1.0,
            "max_tokens": max_tokens,
            "stream": False,
        }
        if response_format is not None:
            payload["response_format"] = response_format

        url = f"{self.base_url}/chat/completions"
        r = requests.post(url, json=payload, timeout=self.timeout)
        r.raise_for_status()
        data = r.json()
        return data["choices"][0]["message"]["content"]

    def _mock_response(self, prompt: str, pass_label: str) -> str:
        # Trivially balanced mock for offline dry runs
        if pass_label == "pass1":
            return json.dumps({
                "Is_Malpractice_Case": "UNKNOWN",
                "Appendicitis_or_Appendectomy_Index_Episode": "UNKNOWN",
                "Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm": "UNKNOWN",
                "likely_case_type": "unclear",
                "full_extraction_warranted": "NO",
                "rationale": "mock",
            })
        else:
            return json.dumps({c: "UNKNOWN" for c in TEMPLATE_COLUMNS_PLACEHOLDER})


# -----------------------------------------------------------------------------
# Prompts
# -----------------------------------------------------------------------------
PASS1_SYSTEM = (
    "You are a medicolegal screening assistant. You read a single Westlaw court opinion and decide, "
    "with a conservative evidence-based approach, whether the case is (a) a malpractice/negligence case, "
    "(b) centered on appendicitis or appendectomy as the clinical index episode, and (c) has a clinically "
    "meaningful appendicitis/appendectomy-related harm. Use YES only when clearly supported. Use NO only "
    "when the opposite is fairly clear. Otherwise UNKNOWN. Respond with clean JSON only."
)


PASS1_USER_TEMPLATE = """Return a JSON object with exactly these keys:

{{
  "Is_Malpractice_Case": "YES | NO | UNKNOWN",
  "Appendicitis_or_Appendectomy_Index_Episode": "YES | NO | UNKNOWN",
  "Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm": "YES | NO | UNKNOWN",
  "likely_case_type": "delayed_diagnosis | operative_complication | postop_management | payment_dispute | incidental_history | unclear",
  "full_extraction_warranted": "YES | NO",
  "rationale": "<=200 chars"
}}

Guidance:
- Delayed/missed diagnosis of appendicitis with meaningful harm IS a true positive even if no appendectomy was performed.
- Operative complications during appendectomy (bowel injury, leak, abscess, stump appendicitis, wrong tissue removed, failure to remove appendix) ARE true positives.
- Incidental "history of appendectomy" mentions in Social Security / disability / ERISA / insurance / prison civil-rights / unrelated PI cases are NOT true positives — set likely_case_type="incidental_history".
- Payment/coverage/billing disputes about appendectomy care are NOT true positives — set likely_case_type="payment_dispute".
- Set full_extraction_warranted="YES" only when at least two of the three screening fields are YES, or when all three are plausibly YES.

Respond with ONLY the JSON object.

CASE TEXT (truncated to ~8k chars):
{case_text}
"""


PASS2_SYSTEM = (
    "You are an expert medical-legal analyst specializing in surgical malpractice litigation, with deep expertise "
    "in abdominal surgery cases including appendectomy procedures and complications. Your task is to perform highly "
    "accurate, structured data extraction from Westlaw full-text court opinions into the exact schema defined below. "
    "You strictly follow the Data_Dictionary and Read_Me rules. You never hallucinate or force values. When information "
    "is unclear, ambiguous, or absent, use \"UNKNOWN\" (or leave blank only when explicitly allowed). You preserve exact "
    "verbatim snippets for key clinical and legal fields to enable manual reviewer verification and auditability.\n\n"
    "DOMAIN-SPECIFIC ADAPTATION FOR APPENDECTOMY CASES:\n"
    "- These are appendectomy (appendix removal) or appendicitis-related cases. The underlying schema is the appendectomy analogue of "
    "the chole project's schema. Treat appendectomy cases on their own terms.\n"
    "- Core_Analytic_Case = YES only if this is a malpractice/negligence case AND (appendectomy is the index procedure OR the lawsuit "
    "centers on a delayed/missed appendicitis diagnosis or delayed operative management) AND there is a clear clinical harm "
    "(perforation, abscess, peritonitis, sepsis, fertility injury, bowel injury, reoperation, stump appendicitis, death, etc.).\n"
    "- Delayed diagnosis of appendicitis counts as Index_Procedure_Appendectomy-adjacent even if surgery was never performed. "
    "Index_Procedure_Appendectomy should reflect whether appendectomy was actually done (YES/NO/UNKNOWN).\n"
    "- Populate Exclusion_Reason clearly when the case is non-malpractice, non-index appendicitis/appendectomy, or otherwise outside analytic scope."
)


PASS2_USER_TEMPLATE = """Extract EVERY field from the following Westlaw case text according to the Data Dictionary below.

DATA DICTIONARY (allowed values and definitions):
{data_dictionary}

CASE TEXT:
{case_text}

EXACT OUTPUT SCHEMA — respond ONLY with valid JSON using these exact keys (do not add, omit, or rename any):
{schema_keys_json}

STRICT RULES:
- Do not force values. Use "UNKNOWN" liberally when truly unclear.
- Settlement_Amount / Damages_Award only when explicitly stated with a numeric value.
- Preserve verbatim snippets in Operative_Text_Snippet, Difficulty_Text_Snippet, Recognition_Text_Snippet, Expert_Criticism_Text, Defense_Strategy_Summary.
- For categorical fields (Difficulty_Assessability, Legal_Outcome, Expert_Testimony_Type, etc.), use ONLY the allowed options.
- Reviewer_Confidence_Score is a 1-5 integer (1=low, 5=high). If the text is thin or uncertain, use 2 or 3 and set Needs_Manual_Review=YES.
- Separate allegation from documented fact, plaintiff expert from defense expert, expert opinion from court finding.
- Do not infer operative difficulty just from a bad outcome.
- Do not assume perforation means negligence.
- Do not assume failed appendectomy means negligence without a supporting theory or expert criticism.

Respond exclusively with valid JSON. No explanations, no markdown.

JSON:
"""


# -----------------------------------------------------------------------------
# Template handling
# -----------------------------------------------------------------------------
def load_template_columns(template_path: Path) -> Tuple[List[str], List[str], List[Tuple[str, str, str, str]]]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed")
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws_master = wb["Case_Master_Template"]
    master_cols = [c for c in next(ws_master.iter_rows(min_row=1, max_row=1, values_only=True)) if c]
    ws_ext = wb["Extended_Extraction"]
    ext_cols = [c for c in next(ws_ext.iter_rows(min_row=1, max_row=1, values_only=True)) if c]
    dd_rows: List[Tuple[str, str, str, str]] = []
    if "Data_Dictionary" in wb.sheetnames:
        ws_dd = wb["Data_Dictionary"]
        for r, row in enumerate(ws_dd.iter_rows(values_only=True)):
            if r == 0:
                continue
            if row and row[0]:
                dd_rows.append((row[0] or "", row[1] or "", row[2] or "", row[3] or ""))
    return master_cols, ext_cols, dd_rows


TEMPLATE_COLUMNS_PLACEHOLDER: List[str] = []  # filled at runtime from Template.xlsx


# -----------------------------------------------------------------------------
# JSON parsing robustness
# -----------------------------------------------------------------------------
def parse_json_strict(s: str) -> Optional[Dict[str, Any]]:
    s = s.strip()
    if s.startswith("```"):
        # strip fenced block if the model slips one in
        s = re.sub(r"^```(?:json)?\s*", "", s)
        s = re.sub(r"\s*```$", "", s)
    try:
        return json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if not m:
            return None
        try:
            return json.loads(m.group(0))
        except Exception:
            return None


# -----------------------------------------------------------------------------
# Extraction orchestration
# -----------------------------------------------------------------------------
def truncate_for_pass1(text: str, cap: int = 8000) -> str:
    if len(text) <= cap:
        return text
    head = text[: int(cap * 0.7)]
    tail = text[-int(cap * 0.3):]
    return head + "\n...[truncated]...\n" + tail


def truncate_for_pass2(text: str, cap: int = 18000) -> str:
    if len(text) <= cap:
        return text
    head = text[: int(cap * 0.7)]
    tail = text[-int(cap * 0.3):]
    return head + "\n...[truncated]...\n" + tail


def pass1_classify(case: CaseChunk, client: LLMClient) -> Dict[str, Any]:
    messages = [
        {"role": "system", "content": PASS1_SYSTEM},
        {"role": "user", "content": PASS1_USER_TEMPLATE.format(case_text=truncate_for_pass1(case.text))},
    ]
    rf = {"type": "json_object"}
    try:
        raw = client.chat(messages, response_format=rf, max_tokens=600, temperature=0.0)
    except Exception as e:
        return {"_error": str(e), "Is_Malpractice_Case": "UNKNOWN",
                "Appendicitis_or_Appendectomy_Index_Episode": "UNKNOWN",
                "Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm": "UNKNOWN",
                "likely_case_type": "unclear", "full_extraction_warranted": "NO",
                "rationale": f"pass1_error: {e}"}
    parsed = parse_json_strict(raw)
    if parsed is None:
        return {"_error": "bad json",
                "Is_Malpractice_Case": "UNKNOWN",
                "Appendicitis_or_Appendectomy_Index_Episode": "UNKNOWN",
                "Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm": "UNKNOWN",
                "likely_case_type": "unclear", "full_extraction_warranted": "NO",
                "rationale": "pass1_parse_fail"}
    return parsed


def data_dict_to_prompt_block(dd_rows: List[Tuple[str, str, str, str]]) -> str:
    lines = []
    for name, defn, allowed, src in dd_rows:
        lines.append(f"- {name}: {defn} (Allowed: {allowed})")
    return "\n".join(lines)


def pass2_extract(case: CaseChunk, client: LLMClient, schema_columns: List[str],
                  data_dict_block: str) -> Dict[str, Any]:
    schema_json = json.dumps(schema_columns, indent=2)
    user = PASS2_USER_TEMPLATE.format(
        data_dictionary=data_dict_block[:12000],
        case_text=truncate_for_pass2(case.text),
        schema_keys_json=schema_json,
    )
    messages = [
        {"role": "system", "content": PASS2_SYSTEM},
        {"role": "user", "content": user},
    ]
    rf = {"type": "json_object"}
    try:
        raw = client.chat(messages, response_format=rf, max_tokens=4500, temperature=0.0)
    except Exception as e:
        return {"_error": str(e)}
    parsed = parse_json_strict(raw)
    if parsed is None:
        return {"_error": "bad_json", "_raw": raw[:2000]}
    return parsed


# -----------------------------------------------------------------------------
# Row assembly, dedup, output
# -----------------------------------------------------------------------------
def derive_core_analytic(rec: Dict[str, Any]) -> Tuple[str, str]:
    A = rec.get("Is_Malpractice_Case", "UNKNOWN")
    B = rec.get("Appendicitis_or_Appendectomy_Index_Episode", "UNKNOWN")
    C = rec.get("Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm", "UNKNOWN")
    if A == "YES" and B == "YES" and C == "YES":
        return "YES", ""
    reasons = []
    if A != "YES":
        reasons.append(f"Is_Malpractice_Case={A}")
    if B != "YES":
        reasons.append(f"Appendicitis_Index={B}")
    if C != "YES":
        reasons.append(f"Harm={C}")
    return "NO", "; ".join(reasons)


def norm_key(s: str) -> str:
    s = (s or "").lower().strip()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def assign_dup_ids(rows: List[Dict[str, Any]]) -> None:
    seen: Dict[str, str] = {}
    counter = 0
    for r in rows:
        k = norm_key(r.get("Case_Name", "")) + "|" + norm_key(r.get("Citation", ""))
        if not k.strip("|"):
            continue
        if k in seen:
            r["Duplicate_Group_ID"] = seen[k]
            if not r.get("Exclusion_Reason"):
                r["Exclusion_Reason"] = "duplicate"
            elif "duplicate" not in r["Exclusion_Reason"]:
                r["Exclusion_Reason"] += "; duplicate"
        else:
            counter += 1
            dup_id = f"DUP-{counter:05d}"
            seen[k] = dup_id


def ensure_all_columns(rec: Dict[str, Any], cols: List[str]) -> Dict[str, Any]:
    out = {}
    for c in cols:
        v = rec.get(c, "")
        if v is None:
            v = ""
        # flatten lists
        if isinstance(v, list):
            v = ", ".join(str(x) for x in v)
        out[c] = v
    return out


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-glob", required=True)
    ap.add_argument("--template", required=True)
    ap.add_argument("--output-xlsx", required=True)
    ap.add_argument("--output-csv")
    ap.add_argument("--output-jsonl")
    ap.add_argument("--output-extended-csv")
    ap.add_argument("--output-manifest")
    ap.add_argument("--output-exclusions")
    ap.add_argument("--base-url", default="http://localhost:8000/v1")
    ap.add_argument("--model", default="gptoss120b")
    ap.add_argument("--workers", type=int, default=4)
    ap.add_argument("--max-cases", type=int, default=0, help="cap total cases for dry runs (0 = all)")
    ap.add_argument("--only-first-pass", action="store_true")
    ap.add_argument("--force-second-pass", action="store_true",
                    help="Run pass2 on every case, not just positives (use for small validation runs)")
    ap.add_argument("--search-group", default="Westlaw appendectomy search")
    ap.add_argument("--search-term", default="appendectomy")
    ap.add_argument("--mock-llm", action="store_true")
    ap.add_argument("--checkpoint-jsonl", default="",
                    help="Optional separate JSONL path for resumable checkpoints (default = --output-jsonl)")
    args = ap.parse_args()

    master_cols, ext_cols, dd_rows = load_template_columns(Path(args.template))
    global TEMPLATE_COLUMNS_PLACEHOLDER
    TEMPLATE_COLUMNS_PLACEHOLDER = list(master_cols)
    data_dict_block = data_dict_to_prompt_block(dd_rows)

    rtf_paths = sorted(Path(p) for p in glob.glob(args.input_glob))
    if not rtf_paths:
        print(f"[ERROR] no RTFs match {args.input_glob}", file=sys.stderr)
        sys.exit(2)
    print(f"Found {len(rtf_paths)} RTF file(s)")

    cases = build_case_chunks(rtf_paths)
    if args.max_cases:
        cases = cases[: args.max_cases]
    print(f"Parsed {len(cases)} case chunks")

    # Manifest
    if args.output_manifest:
        Path(args.output_manifest).parent.mkdir(parents=True, exist_ok=True)
        with open(args.output_manifest, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["Search_ID", "File_Name", "Segment_Index",
                                              "Case_Name_Hint", "Citation_Hint", "Year_Hint",
                                              "Char_Count"])
            w.writeheader()
            for c in cases:
                w.writerow(c.to_manifest_row())
        print(f"Wrote manifest {args.output_manifest}")

    client = LLMClient(args.base_url, args.model, mock=args.mock_llm)

    # Checkpoint-aware resumability
    ckpt_path = Path(args.checkpoint_jsonl or args.output_jsonl or "/tmp/appendectomy_ckpt.jsonl")
    done_ids: set = set()
    if ckpt_path.exists():
        with ckpt_path.open("r", encoding="utf-8") as f:
            for line in f:
                try:
                    done_ids.add(json.loads(line)["Search_ID"])
                except Exception:
                    continue
        print(f"Resuming: {len(done_ids)} cases already checkpointed")

    ckpt_path.parent.mkdir(parents=True, exist_ok=True)
    ckpt_f = ckpt_path.open("a", encoding="utf-8")

    # -----------------------------------------------------------------
    # First pass
    # -----------------------------------------------------------------
    todo = [c for c in cases if c.search_id not in done_ids]
    print(f"Running pass-1 on {len(todo)} cases with {args.workers} workers")
    pass1_results: Dict[str, Dict[str, Any]] = {}
    t0 = time.time()
    with cf.ThreadPoolExecutor(max_workers=args.workers) as pool:
        futs = {pool.submit(pass1_classify, c, client): c for c in todo}
        for i, fut in enumerate(cf.as_completed(futs), start=1):
            c = futs[fut]
            try:
                pass1_results[c.search_id] = fut.result()
            except Exception as e:
                pass1_results[c.search_id] = {"_error": str(e)}
            if i % 20 == 0 or i == len(todo):
                print(f"  pass1 {i}/{len(todo)} elapsed={time.time()-t0:.1f}s")

    # -----------------------------------------------------------------
    # Second pass (only on likely positives unless --force-second-pass)
    # -----------------------------------------------------------------
    def is_positive(p1: Dict[str, Any]) -> bool:
        if p1.get("full_extraction_warranted") == "YES":
            return True
        pos = sum(1 for k in ("Is_Malpractice_Case",
                              "Appendicitis_or_Appendectomy_Index_Episode",
                              "Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm")
                  if p1.get(k) == "YES")
        return pos >= 2

    cand: List[CaseChunk] = []
    for c in todo:
        p1 = pass1_results.get(c.search_id, {})
        if args.only_first_pass:
            continue
        if args.force_second_pass or is_positive(p1):
            cand.append(c)
    print(f"Running pass-2 on {len(cand)} candidates")

    pass2_results: Dict[str, Dict[str, Any]] = {}
    if cand:
        with cf.ThreadPoolExecutor(max_workers=args.workers) as pool:
            futs = {pool.submit(pass2_extract, c, client, master_cols, data_dict_block): c for c in cand}
            for i, fut in enumerate(cf.as_completed(futs), start=1):
                c = futs[fut]
                try:
                    pass2_results[c.search_id] = fut.result()
                except Exception as e:
                    pass2_results[c.search_id] = {"_error": str(e)}
                if i % 10 == 0 or i == len(cand):
                    print(f"  pass2 {i}/{len(cand)} elapsed={time.time()-t0:.1f}s")

    # -----------------------------------------------------------------
    # Assemble rows
    # -----------------------------------------------------------------
    master_rows: List[Dict[str, Any]] = []
    ext_rows: List[Dict[str, Any]] = []
    for c in cases:
        p1 = pass1_results.get(c.search_id, {}) if c.search_id in pass1_results else {}
        p2 = pass2_results.get(c.search_id, {})

        # Deterministic fields — never let the LLM overwrite these.
        DETERMINISTIC_FIELDS = {
            "Search_ID", "Search_Group", "Search_Term", "Source_Database", "File_Name",
        }
        rec: Dict[str, Any] = {}

        # prefer pass2 values, then pass1, then hints
        if p2 and "_error" not in p2:
            for k, v in p2.items():
                if k in master_cols and k not in DETERMINISTIC_FIELDS:
                    rec[k] = v
            rec["Full_Extraction_Performed"] = "YES"
            rec["LLM_Status"] = p2.get("LLM_Status", "full")
        else:
            rec["Full_Extraction_Performed"] = "NO"
            rec["LLM_Status"] = "first_pass_only" if p1 else "skipped"

        # Apply deterministic fields AFTER the LLM merge so they always win.
        rec["Search_ID"] = c.search_id
        rec["Search_Group"] = args.search_group
        rec["Search_Term"] = args.search_term
        rec["Source_Database"] = SOURCE_DATABASE
        rec["File_Name"] = c.file_name

        if p1:
            for k in ("Is_Malpractice_Case",
                      "Appendicitis_or_Appendectomy_Index_Episode",
                      "Has_Clinically_Meaningful_Appendicitis_or_Appendectomy_Harm"):
                rec.setdefault(k, p1.get(k, "UNKNOWN"))
            rec["First_Pass_Likely_Case_Type"] = p1.get("likely_case_type", "unclear")
            rec["First_Pass_Rationale"] = p1.get("rationale", "")

        # hint fallbacks
        if not rec.get("Case_Name"):
            rec["Case_Name"] = c.case_name_hint
        if not rec.get("Citation"):
            rec["Citation"] = c.citation_hint
        if not rec.get("Year"):
            rec["Year"] = c.year_hint

        core, excl = derive_core_analytic(rec)
        rec["Core_Analytic_Case"] = core
        if rec.get("Exclusion_Reason"):
            if excl and excl not in rec["Exclusion_Reason"]:
                rec["Exclusion_Reason"] += "; " + excl
        else:
            rec["Exclusion_Reason"] = excl

        # confidence default
        if rec.get("Reviewer_Confidence_Score") in (None, ""):
            rec["Reviewer_Confidence_Score"] = ""
        try:
            conf = int(rec.get("Reviewer_Confidence_Score") or 0)
        except Exception:
            conf = 0
        rec["Needs_Manual_Review"] = "YES" if (conf and conf <= 3) or rec["LLM_Status"] != "full" else "NO"

        master_rows.append(ensure_all_columns(rec, master_cols))

        # Extended row
        erec = {c2: p2.get(c2, "") if isinstance(p2, dict) else "" for c2 in ext_cols}
        erec["Search_ID"] = c.search_id
        erec["LLM_Status"] = rec["LLM_Status"]
        ext_rows.append(erec)

        # checkpoint
        ckpt_f.write(json.dumps({
            "Search_ID": c.search_id,
            "File_Name": c.file_name,
            "pass1": p1,
            "pass2": p2,
            "final": rec,
        }, ensure_ascii=False) + "\n")
        ckpt_f.flush()

    ckpt_f.close()

    # Dedup
    assign_dup_ids(master_rows)

    # Write Excel — populate the template
    wb = openpyxl.load_workbook(args.template)
    ws = wb["Case_Master_Template"]
    # clear any pre-existing data rows
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)
    for r, rec in enumerate(master_rows, start=2):
        for j, col in enumerate(master_cols, start=1):
            ws.cell(row=r, column=j, value=rec.get(col, ""))
    # Extended sheet
    ws_ext = wb["Extended_Extraction"]
    if ws_ext.max_row > 1:
        ws_ext.delete_rows(2, ws_ext.max_row)
    for r, erec in enumerate(ext_rows, start=2):
        for j, col in enumerate(ext_cols, start=1):
            ws_ext.cell(row=r, column=j, value=erec.get(col, ""))
    Path(args.output_xlsx).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output_xlsx)
    print(f"Wrote xlsx {args.output_xlsx}")

    # CSV
    if args.output_csv:
        with open(args.output_csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=master_cols)
            w.writeheader()
            w.writerows(master_rows)
        print(f"Wrote master csv {args.output_csv}")

    # Extended CSV
    if args.output_extended_csv:
        with open(args.output_extended_csv, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=ext_cols)
            w.writeheader()
            w.writerows(ext_rows)
        print(f"Wrote extended csv {args.output_extended_csv}")

    # Exclusions
    if args.output_exclusions:
        with open(args.output_exclusions, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Search_ID", "Case_Name", "Citation", "Year", "Exclusion_Reason", "First_Pass_Likely_Case_Type"])
            for r in master_rows:
                if r.get("Core_Analytic_Case") != "YES":
                    w.writerow([r.get("Search_ID"), r.get("Case_Name"), r.get("Citation"),
                                r.get("Year"), r.get("Exclusion_Reason"),
                                r.get("First_Pass_Likely_Case_Type")])
        print(f"Wrote exclusions csv {args.output_exclusions}")

    # Summary
    n_total = len(master_rows)
    n_core = sum(1 for r in master_rows if r.get("Core_Analytic_Case") == "YES")
    n_excl = n_total - n_core
    top_reasons: Dict[str, int] = {}
    for r in master_rows:
        if r.get("Core_Analytic_Case") != "YES":
            reason = (r.get("Exclusion_Reason") or "unknown").split(";")[0].strip()
            top_reasons[reason] = top_reasons.get(reason, 0) + 1
    print("\n==================== SUMMARY ====================")
    print(f"Total parsed cases:        {n_total}")
    print(f"Core analytic cases:       {n_core}")
    print(f"Excluded / non-core:       {n_excl}")
    print("Top exclusion reasons:")
    for k, v in sorted(top_reasons.items(), key=lambda kv: -kv[1])[:10]:
        print(f"  {v:>5}  {k}")


if __name__ == "__main__":
    main()
